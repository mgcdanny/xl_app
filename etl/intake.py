
from openpyxl import Workbook,load_workbook
from collections import namedtuple
from datetime import datetime
import xlrd
from xlwt.Utils import cell_to_rowcol2 

def json_serial(obj):
    """JSON serializer for objects not serializable by default json code
    	flask.jsonify can serialize datetime so this function might
    	not be needed"""
    if isinstance(obj, datetime):
        serial = obj.isoformat()
    return serial


def getData(fp):
	"""get data out of excel file based on config file"""
	Data = namedtuple('Data',['loc','type','name'])	
	#data_only takes values of formulas
	wb = load_workbook(filename = fp, data_only=True)
	ws_config = wb.get_sheet_by_name("config")
	ws_form = wb.get_sheet_by_name(wb.get_sheet_names()[0])
	#skip the header / get the rows of data
	rows = ws_config.rows[1:]
	#convert rows to Data
	container = [Data._make([cell.value for cell in row]) for row in rows]
	result = {data.name: ws_form.cell(data.loc.upper()).value for data in container}
	return result

def writeData(fp, outFp):
	"""test of prepopulating excel files with data from other source"""
	wb = load_workbook(filename = fp, data_only=True)
	ws_form = wb.get_sheet_by_name(wb.get_sheet_names()[0])
	wb.save(outFp)

def getData_xlrd(fp):
	Data = namedtuple('Data',['loc','type','name'])
	wb = xlrd.open_workbook(fp)
	ws_config = wb.sheet_by_name('config')
	ws_form = wb.sheet_by_index(0)
	container = [Data._make([c.value for c in ws_config.row(curr_row)]) for curr_row in range(1,ws_config.nrows)]
	result = {data.name: ws_form.cell(*cell_to_rowcol2(data.loc)).value for data in container}
	return result

if __name__ == "__main__":
	print("hello")
	#wb = load_workbook(filename = "Intake_simple.xlsx", data_only=True)
	#writeData("Intake_simple.xlsx", "Intake_simple_test_out.xlsx")
	#wb, container, result = getData_xlrd(fp = "../forms/intake/999")